"""Excel 기반 Repository 구현체 (Phase 1)"""
import uuid
import json
import threading
import pandas as pd
import openpyxl
from datetime import datetime, timezone
from typing import Optional
from pathlib import Path

from app.repositories.base_repository import BaseUserRepository, BaseContentRepository
from app.models.user import UserInDB
from app.models.content import Content, ContentCreate
from app.core.config import settings

# Excel 파일 동시성 제어
_excel_lock = threading.Lock()


class ExcelUserRepository(BaseUserRepository):
    """Excel 기반 사용자 Repository"""

    def __init__(self):
        self._path = settings.EXCEL_PATH

    def _load(self, sheet: str) -> pd.DataFrame:
        return pd.read_excel(self._path, sheet_name=sheet, dtype=str)

    def _save(self, sheet: str, df: pd.DataFrame):
        with _excel_lock:
            wb = openpyxl.load_workbook(self._path)
            if sheet in wb.sheetnames:
                del wb[sheet]
            ws = wb.create_sheet(sheet)
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            wb.save(self._path)

    def get_user_by_emp_id(self, emp_id: str) -> Optional[UserInDB]:
        df = self._load("users")
        row = df[df["emp_id"] == emp_id]
        if row.empty:
            return None
        r = row.iloc[0]
        return UserInDB(
            emp_id=r["emp_id"],
            name=r["name"],
            email=r["email"],
            is_admin=(str(r.get("is_admin", "False")).upper() == "TRUE"),
            is_active=(str(r.get("is_active", "True")).upper() == "TRUE"),
        )

    def list_all_users(self) -> list[UserInDB]:
        df = self._load("users")
        users = []
        for _, r in df.iterrows():
            users.append(UserInDB(
                emp_id=r["emp_id"],
                name=r["name"],
                email=r["email"],
                is_admin=(str(r.get("is_admin", "False")).upper() == "TRUE"),
                is_active=(str(r.get("is_active", "True")).upper() == "TRUE"),
            ))
        return users


class ExcelContentRepository(BaseContentRepository):
    """Excel 기반 컨텐츠 Repository"""

    def __init__(self):
        self._path = settings.EXCEL_PATH

    def _load(self, sheet: str) -> pd.DataFrame:
        try:
            return pd.read_excel(self._path, sheet_name=sheet, dtype=str)
        except ValueError:
            # 시트가 없는 경우 빈 DataFrame 반환
            return pd.DataFrame()

    def _save(self, sheet: str, df: pd.DataFrame):
        with _excel_lock:
            wb = openpyxl.load_workbook(self._path)
            if sheet in wb.sheetnames:
                del wb[sheet]
            ws = wb.create_sheet(sheet)
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            wb.save(self._path)

    def _parse_template_config(self, raw) -> dict:
        """template_config JSON 문자열을 dict로 파싱"""
        try:
            if raw and str(raw) != "nan":
                return json.loads(str(raw))
        except (json.JSONDecodeError, TypeError):
            pass
        return {}

    def get_contents_by_emp_id(self, emp_id: str) -> list[Content]:
        contents_df = self._load("contents")
        if contents_df.empty:
            return []

        confirmations_df = self._load("confirmations")

        user_contents = contents_df[contents_df["target_emp_id"] == emp_id]
        result = []
        for _, r in user_contents.iterrows():
            is_confirmed = False
            confirmed_at = None
            confirmed_text = ""

            if not confirmations_df.empty:
                conf = confirmations_df[
                    (confirmations_df["content_id"] == r["content_id"]) &
                    (confirmations_df["emp_id"] == emp_id)
                ]
                if not conf.empty:
                    is_confirmed = str(conf.iloc[0].get("is_confirmed", "False")).upper() == "TRUE"
                    confirmed_at_str = conf.iloc[0].get("confirmed_at")
                    if confirmed_at_str and str(confirmed_at_str) != "nan":
                        try:
                            confirmed_at = datetime.fromisoformat(str(confirmed_at_str))
                        except (ValueError, TypeError):
                            confirmed_at = None
                    confirmed_text_val = conf.iloc[0].get("confirmed_text", "")
                    if confirmed_text_val and str(confirmed_text_val) != "nan":
                        confirmed_text = str(confirmed_text_val)

            # template_config JSON 파싱
            template_config = self._parse_template_config(r.get("template_config"))

            result.append(Content(
                content_id=r["content_id"],
                title=r["title"],
                target_emp_id=r["target_emp_id"],
                body_1=r.get("body_1", ""),
                body_2=r.get("body_2") if str(r.get("body_2", "nan")) != "nan" else None,
                body_3=r.get("body_3") if str(r.get("body_3", "nan")) != "nan" else None,
                answer_type=r.get("answer_type", "yes_no") if str(r.get("answer_type", "nan")) != "nan" else "yes_no",
                template_config=template_config,
                file_path=r.get("file_path") if str(r.get("file_path", "nan")) != "nan" and r.get("file_path") else None,
                created_by=r.get("created_by", ""),
                is_confirmed=is_confirmed,
                confirmed_at=confirmed_at,
                confirmed_text=confirmed_text,
            ))
        return result

    def get_all_contents(self) -> list[Content]:
        contents_df = self._load("contents")
        if contents_df.empty:
            return []

        confirmations_df = self._load("confirmations")
        result = []
        for _, r in contents_df.iterrows():
            is_confirmed = False
            if not confirmations_df.empty:
                conf = confirmations_df[
                    (confirmations_df["content_id"] == r["content_id"])
                ]
                if not conf.empty:
                    is_confirmed = str(conf.iloc[0].get("is_confirmed", "False")).upper() == "TRUE"

            template_config = self._parse_template_config(r.get("template_config"))

            result.append(Content(
                content_id=r["content_id"],
                title=r["title"],
                target_emp_id=r["target_emp_id"],
                body_1=r.get("body_1", ""),
                body_2=r.get("body_2") if str(r.get("body_2", "nan")) != "nan" else None,
                body_3=r.get("body_3") if str(r.get("body_3", "nan")) != "nan" else None,
                answer_type=r.get("answer_type", "yes_no") if str(r.get("answer_type", "nan")) != "nan" else "yes_no",
                template_config=template_config,
                file_path=r.get("file_path") if str(r.get("file_path", "nan")) != "nan" and r.get("file_path") else None,
                created_by=r.get("created_by", ""),
                is_confirmed=is_confirmed,
            ))
        return result

    def create_content(self, payload: ContentCreate, creator_id: str) -> Content:
        contents_df = self._load("contents")
        content_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()

        new_row = {
            "content_id": content_id,
            "title": payload.title,
            "target_emp_id": payload.target_emp_id,
            "body_1": payload.body_1,
            "body_2": payload.body_2 or "",
            "body_3": payload.body_3 or "",
            "answer_type": payload.answer_type,
            "template_config": json.dumps(payload.template_config, ensure_ascii=False),
            "file_path": "",          # 파일 업로드는 별도 엔드포인트 처리
            "created_by": creator_id,
            "created_at": now,
            "updated_at": "",
        }

        if contents_df.empty:
            contents_df = pd.DataFrame([new_row])
        else:
            contents_df = pd.concat([contents_df, pd.DataFrame([new_row])], ignore_index=True)

        self._save("contents", contents_df)

        # confirmations에 초기 레코드 생성
        confirmations_df = self._load("confirmations")
        conf_row = {
            "confirmation_id": str(uuid.uuid4()),
            "content_id": content_id,
            "emp_id": payload.target_emp_id,
            "is_confirmed": "False",
            "confirmed_at": "",
            "confirmed_text": "",
        }

        if confirmations_df.empty:
            confirmations_df = pd.DataFrame([conf_row])
        else:
            confirmations_df = pd.concat([confirmations_df, pd.DataFrame([conf_row])], ignore_index=True)

        self._save("confirmations", confirmations_df)

        return Content(
            content_id=content_id,
            title=payload.title,
            target_emp_id=payload.target_emp_id,
            body_1=payload.body_1,
            body_2=payload.body_2,
            body_3=payload.body_3,
            answer_type=payload.answer_type,
            template_config=payload.template_config,
            created_by=creator_id,
        )

    def update_content(self, content_id: str, payload: dict) -> Optional[Content]:
        contents_df = self._load("contents")
        if contents_df.empty:
            return None

        idx = contents_df[contents_df["content_id"] == content_id].index
        if idx.empty:
            return None

        row_idx = idx[0]
        for key, value in payload.items():
            if value is not None and key in contents_df.columns:
                contents_df.at[row_idx, key] = value

        contents_df.at[row_idx, "updated_at"] = datetime.now(timezone.utc).isoformat()
        self._save("contents", contents_df)

        r = contents_df.iloc[row_idx]
        template_config = self._parse_template_config(r.get("template_config"))

        return Content(
            content_id=r["content_id"],
            title=r["title"],
            target_emp_id=r["target_emp_id"],
            body_1=r.get("body_1", ""),
            body_2=r.get("body_2") if str(r.get("body_2", "nan")) != "nan" else None,
            body_3=r.get("body_3") if str(r.get("body_3", "nan")) != "nan" else None,
            answer_type=r.get("answer_type", "yes_no") if str(r.get("answer_type", "nan")) != "nan" else "yes_no",
            template_config=template_config,
            created_by=r.get("created_by", ""),
        )

    def delete_content(self, content_id: str) -> bool:
        df = self._load("contents")
        mask = df["content_id"] == content_id
        if not mask.any():
            return False
        df = df[~mask].reset_index(drop=True)
        self._save("contents", df)

        # 연관 확인 기록도 함께 삭제
        df_conf = self._load("confirmations")
        df_conf = df_conf[df_conf["content_id"] != content_id].reset_index(drop=True)
        self._save("confirmations", df_conf)

        return True

    def set_confirmation(self, content_id: str, emp_id: str, confirmed: bool, confirmed_text: str = "") -> bool:
        confirmations_df = self._load("confirmations")
        if confirmations_df.empty:
            return False

        mask = (confirmations_df["content_id"] == content_id) & (confirmations_df["emp_id"] == emp_id)
        matching = confirmations_df[mask]

        if matching.empty:
            # 레코드가 없으면 새로 생성
            conf_row = {
                "confirmation_id": str(uuid.uuid4()),
                "content_id": content_id,
                "emp_id": emp_id,
                "is_confirmed": str(confirmed).upper(),
                "confirmed_at": datetime.now(timezone.utc).isoformat() if confirmed else "",
                "confirmed_text": confirmed_text,
            }
            confirmations_df = pd.concat([confirmations_df, pd.DataFrame([conf_row])], ignore_index=True)
        else:
            idx = matching.index[0]
            confirmations_df.at[idx, "is_confirmed"] = str(confirmed).upper()
            confirmations_df.at[idx, "confirmed_at"] = datetime.now(timezone.utc).isoformat() if confirmed else ""
            confirmations_df.at[idx, "confirmed_text"] = confirmed_text

        self._save("confirmations", confirmations_df)
        return True

    def get_all_confirmations(self) -> list[dict]:
        confirmations_df = self._load("confirmations")
        if confirmations_df.empty:
            return []
        confirmations_df = confirmations_df.fillna("")
        return confirmations_df.to_dict("records")

    def get_confirmations_filtered(
        self,
        content_id: str | None = None,
        emp_id: str | None = None
    ) -> list[dict]:
        df_conf = self._load("confirmations")
        df_cont = self._load("contents")
        df_user = self._load("users")

        if df_conf.empty:
            return []

        # 컨텐츠 제목, 사용자 이름 조인
        merged = df_conf
        if not df_cont.empty and "title" in df_cont.columns:
            merged = merged.merge(df_cont[["content_id", "title"]], on="content_id", how="left")
        else:
            merged["title"] = ""
        if not df_user.empty and "name" in df_user.columns:
            merged = merged.merge(df_user[["emp_id", "name"]], on="emp_id", how="left")
        else:
            merged["name"] = ""

        if content_id:
            merged = merged[merged["content_id"] == content_id]
        if emp_id:
            merged = merged[merged["emp_id"] == emp_id]

        result_cols = ["content_id", "title", "emp_id", "name", "is_confirmed", "confirmed_at"]
        for col in result_cols:
            if col not in merged.columns:
                merged[col] = ""

        merged = merged[result_cols].fillna("")
        return merged.to_dict(orient="records")

    # ── PATCH 2: 추가 메서드 ──

    def get_content_by_id(self, content_id: str) -> dict | None:
        """단일 컨텐츠 조회 (dict 반환)"""
        df = self._load("contents")
        row = df[df["content_id"] == content_id]
        if row.empty:
            return None
        r = row.iloc[0].to_dict()
        r["template_config"] = self._parse_template_config(r.get("template_config"))
        return r

    def update_file_path(self, content_id: str, file_path: str):
        """교육자료 파일 경로 업데이트"""
        with _excel_lock:
            df = self._load("contents")
            df.loc[df["content_id"] == content_id, "file_path"] = file_path
            self._save("contents", df)
