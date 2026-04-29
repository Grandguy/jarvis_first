"""초기 Excel 데이터 파일 생성 스크립트"""
import openpyxl
from datetime import datetime, timezone

wb = openpyxl.Workbook()

# Sheet 1: users
ws_users = wb.active
ws_users.title = "users"
ws_users.append(["emp_id", "name", "email", "is_admin", "is_active", "created_at"])
ws_users.append(["EMP001", "김관리자", "admin@example.com", "True", "True", datetime.now(timezone.utc).isoformat()])
ws_users.append(["EMP002", "이사원", "user1@example.com", "False", "True", datetime.now(timezone.utc).isoformat()])
ws_users.append(["EMP003", "박대리", "user2@example.com", "False", "True", datetime.now(timezone.utc).isoformat()])

# Sheet 2: contents
ws_contents = wb.create_sheet("contents")
ws_contents.append(["content_id", "title", "target_emp_id", "body_1", "body_2", "body_3", "created_by", "created_at", "updated_at"])
ws_contents.append([
    "sample-001", "4월 업무 공지", "EMP002",
    "4월 정기 보안 교육이 예정되어 있습니다. 반드시 참석해주세요.",
    "교육 일시: 2026년 4월 30일 14:00",
    "", "EMP001",
    datetime.now(timezone.utc).isoformat(), ""
])
ws_contents.append([
    "sample-002", "신규 시스템 안내", "EMP003",
    "새로운 컨텐츠 열람 시스템이 도입되었습니다.",
    "사번과 OTP를 통해 본인 확인 후 열람 가능합니다.",
    "문의사항은 관리자에게 연락해주세요.", "EMP001",
    datetime.now(timezone.utc).isoformat(), ""
])

# Sheet 3: confirmations
ws_confirmations = wb.create_sheet("confirmations")
ws_confirmations.append(["confirmation_id", "content_id", "emp_id", "is_confirmed", "confirmed_at"])
ws_confirmations.append(["conf-001", "sample-001", "EMP002", "False", ""])
ws_confirmations.append(["conf-002", "sample-002", "EMP003", "False", ""])

wb.save("data/master.xlsx")
print("✅ data/master.xlsx 생성 완료!")
print("   - users: 3명 (관리자 1, 사용자 2)")
print("   - contents: 2건")
print("   - confirmations: 2건")
